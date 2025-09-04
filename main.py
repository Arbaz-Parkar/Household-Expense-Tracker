import os
from typing import Tuple, Dict, Any
import pandas as pd
import seaborn as sns
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage


# 1) Excel File Reading
def load_data(file_path: str) -> pd.DataFrame:
    """Read household expenses dataset from Excel into a DataFrame."""
    df = pd.read_excel(file_path)
    return df


# 2) Data Cleaning
def _amount_is_valid(value: Any) -> bool:
    """Check if expense amount is within a realistic range (100–50,000)."""
    try:
        v = int(value)
        return 100 <= v <= 50000
    except:
        return False


def _payment_mode_is_valid(value: Any) -> bool:
    """Check if payment mode is one of the expected categories."""
    valid_modes = {"Cash", "UPI", "Card", "NetBanking"}
    return str(value).strip() in valid_modes


def clean_data(df: pd.DataFrame) -> pd.DataFrame:
    """Clean dataset: handle missing values, duplicates, types, and validations."""
    # Standardize column names
    df.columns = df.columns.str.strip().str.replace(" ", "_").str.lower()

    # Fill missing values
    df = df.fillna({
        "notes": "No Notes",
        "amount": 0,
        "payment_mode": "Unknown",
        "category": "Misc"
    })

    # Remove duplicates
    df = df.drop_duplicates()

    # Convert amount to int
    if "amount" in df.columns:
        df["amount"] = pd.to_numeric(df["amount"], errors="coerce").fillna(0).astype(int)

    # Ensure date is datetime
    if "date" in df.columns:
        df["date"] = pd.to_datetime(df["date"], errors="coerce")

    # Validation flags
    df["amount_valid"] = df["amount"].apply(_amount_is_valid)
    df["payment_mode_valid"] = df["payment_mode"].apply(_payment_mode_is_valid)

    return df


# 3) Data Analysis
def analyze_data(df: pd.DataFrame) -> Dict[str, Any]:
    """Perform insights like mean, max, grouping, counts, filtering."""
    results = {}

    # Average, Max, Min
    results["average_expense"] = df["amount"].mean()
    results["max_expense"] = df["amount"].max()
    results["min_expense"] = df["amount"].min()

    # Group by category
    results["category_totals"] = df.groupby("category")["amount"].sum().sort_values(ascending=False)

    # Category averages
    results["category_averages"] = df.groupby("category")["amount"].mean().sort_values(ascending=False)

    # Payment mode counts
    results["payment_counts"] = df["payment_mode"].value_counts()

    # Monthly totals
    if "date" in df.columns:
        df["month"] = df["date"].dt.to_period("M")
        results["monthly_totals"] = df.groupby("month")["amount"].sum()
    else:
        results["monthly_totals"] = pd.Series()

    # Top 5 expensive items
    results["top5_items"] = df.sort_values(by="amount", ascending=False).head(5)

    # Filter > 5000
    results["above_5000"] = df[df["amount"] > 5000]

    # Sort all expenses
    results["sorted_expenses"] = df.sort_values(by="amount", ascending=False)

    return results


# 4) Report Generation
def export_report(output_path: str, results: Dict[str, Any]) -> None:
    """Save analysis results into an Excel file with multiple sheets."""
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Summary
        pd.DataFrame({
            "Average Expense": [results["average_expense"]],
            "Max Expense": [results["max_expense"]],
            "Min Expense": [results["min_expense"]]
        }).to_excel(writer, sheet_name="Summary", index=False)

        # Category totals
        results["category_totals"].rename_axis("Category").reset_index(name="Total Amount") \
            .to_excel(writer, sheet_name="Category Totals", index=False)

        # Category averages
        results["category_averages"].rename_axis("Category").reset_index(name="Avg Amount") \
            .to_excel(writer, sheet_name="Category Averages", index=False)

        # Payment modes
        results["payment_counts"].rename_axis("Payment Mode").reset_index(name="Count") \
            .to_excel(writer, sheet_name="Payment Modes", index=False)

        # Monthly totals
        results["monthly_totals"].rename_axis("Month").reset_index(name="Total Amount") \
            .to_excel(writer, sheet_name="Monthly Totals", index=False)

        # Top 5 items
        results["top5_items"].to_excel(writer, sheet_name="Top 5 Items", index=False)

        # Above 5000
        results["above_5000"].to_excel(writer, sheet_name="Above 5000", index=False)

        # Sorted
        results["sorted_expenses"].to_excel(writer, sheet_name="Sorted Expenses", index=False)

        # Charts placeholder
        pd.DataFrame({"Charts": ["See embedded images"]}).to_excel(writer, sheet_name="Charts", index=False)


# 5) Charts & Visualization
def generate_charts(df: pd.DataFrame, charts_dir: str) -> Dict[str, str]:
    os.makedirs(charts_dir, exist_ok=True)
    paths: Dict[str, str] = {}
    sns.set_theme(style="whitegrid")

    # Bar: Expenses by Category
    plt.figure(figsize=(8, 5))
    totals = df.groupby("category")["amount"].sum().sort_values(ascending=False)
    sns.barplot(x=totals.index, y=totals.values, palette="Blues_d")
    plt.title("Expenses by Category")
    plt.xlabel("Category")
    plt.ylabel("Total Amount")
    cat_path = os.path.join(charts_dir, "category_expenses_bar.png")
    plt.tight_layout()
    plt.savefig(cat_path, dpi=150)
    plt.close()
    paths["category_bar"] = cat_path

    # Pie: Payment Mode Distribution
    plt.figure(figsize=(6, 6))
    df["payment_mode"].value_counts().plot(kind="pie", autopct="%1.1f%%", startangle=90)
    plt.title("Payment Mode Distribution")
    plt.ylabel("")
    pie_path = os.path.join(charts_dir, "payment_mode_pie.png")
    plt.tight_layout()
    plt.savefig(pie_path, dpi=150)
    plt.close()
    paths["payment_pie"] = pie_path

    # Line: Monthly Totals
    if "date" in df.columns:
        plt.figure(figsize=(10, 5))
        monthly = df.groupby(df["date"].dt.to_period("M"))["amount"].sum()
        plt.plot(monthly.index.astype(str), monthly.values, marker="o")
        plt.title("Monthly Expenses Over Time")
        plt.xlabel("Month")
        plt.ylabel("Total Amount")
        line_path = os.path.join(charts_dir, "monthly_expenses_line.png")
        plt.tight_layout()
        plt.savefig(line_path, dpi=150)
        plt.close()
        paths["monthly_line"] = line_path

    # Histogram: Expense Distribution
    plt.figure(figsize=(8, 5))
    sns.histplot(df["amount"], bins=15, kde=True, color="#4C72B0")
    plt.title("Expense Distribution")
    plt.xlabel("Amount")
    plt.ylabel("Frequency")
    hist_path = os.path.join(charts_dir, "expense_hist.png")
    plt.tight_layout()
    plt.savefig(hist_path, dpi=150)
    plt.close()
    paths["histogram"] = hist_path

    return paths


def embed_charts_in_excel(excel_path: str, chart_paths: Dict[str, str]) -> None:
    wb = load_workbook(excel_path)
    if "Charts" not in wb.sheetnames:
        wb.create_sheet("Charts")
    ws = wb["Charts"]

    row = 1
    for label, path in chart_paths.items():
        if os.path.exists(path):
            img = XLImage(path)
            cell = f"A{row}"
            ws.add_image(img, cell)
            row += 20
    wb.save(excel_path)


# 6) Orchestration
def main():
    source_file = "household_expenses.xlsx"
    output_excel = "cleaned_expenses.xlsx"
    charts_dir = "charts"

    # Load
    df = load_data(source_file)

    # Clean
    df_clean = clean_data(df)

    # Analyze
    results = analyze_data(df_clean)

    # Export report
    export_report(output_excel, results)

    # Charts
    chart_paths = generate_charts(df_clean, charts_dir)
    embed_charts_in_excel(output_excel, chart_paths)

    # Console summary
    print("\nExpense Analysis Report")
    print("=======================")
    print(f"Average expense: {results['average_expense']:.2f}")
    print(f"Max expense: {results['max_expense']}")
    print(f"Min expense: {results['min_expense']}")
    print("\nTotal by category:")
    print(results["category_totals"])
    print("\nPayment mode distribution:")
    print(results["payment_counts"])
    print("\nTop 5 expensive items:")
    print(results["top5_items"][['date', 'category', 'item', 'amount']])
    print(f"\nReport written to: {output_excel}")


if __name__ == "__main__":
    main()
